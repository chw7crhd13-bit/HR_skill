#!/usr/bin/env python3
"""Create a HR compensation Excel demo workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError as exc:
    raise SystemExit("Missing dependency: openpyxl. Run `python3 -m pip install openpyxl`.") from exc

RAW_HEADERS = [
    "month", "employee_id", "employee_name", "department", "city", "grade", "status",
    "base_salary", "allowance", "bonus", "social_insurance", "housing_fund", "tax",
]

RAW_ROWS = [
    ["2026-04", "E1001", "Chen Yi", "HR", "Shanghai", "G6", "Active", 12000, 800, 1500, 1260, 840, 980],
    ["2026-04", "E1002", "Li Na", "Finance", "Shanghai", "G7", "Active", 16800, 1200, 2200, 1764, 1176, 1680],
    ["2026-04", "E1003", "Wang Lei", "Sales", "Hangzhou", "G5", "Active", 9800, 600, 1100, 1029, 686, 620],
    ["2026-04", "E1004", "Zhao Min", "Tech", "Shenzhen", "G8", "Active", 23500, 1500, 3800, 2468, 1645, 3100],
    ["2026-04", "E1005", "Sun Ke", "HR", "Beijing", "G4", "Inactive", 7200, 400, 0, 756, 504, 280],
    ["2026-04", "E1006", "He Lin", "Ops", "Shanghai", "G6", "Active", 11800, 700, 900, 1239, 826, 820],
    ["2026-04", "E1007", "Qian Yu", "Finance", "Beijing", "G5", "Active", 10500, 600, 1300, 1103, 735, 710],
    ["2026-04", "E1008", "Liu Fang", "Tech", "Shenzhen", "G7", "Active", 18200, 1300, 2800, 1911, 1274, 2050],
    ["2026-04", "E1009", "Zhou Rui", "Sales", "Hangzhou", "G6", "Active", 12100, 900, 2100, 1271, 847, 1080],
    ["2026-04", "E1010", "Wu Xin", "HR", "Shanghai", "G5", "Active", 10300, 600, 1000, 1082, 721, 650],
    ["2026-04", "E1010", "Wu Xin", "HR", "Shanghai", "G5", "Active", 10300, 600, 1000, 1082, 721, 650],
    ["2026-04", "", "Missing ID", "Finance", "Shanghai", "G6", "Active", 11900, 700, 800, 1250, 833, 760],
    ["2026-04", "E1012", "Band Risk", "Tech", "Shenzhen", "G4", "Active", 22000, 900, 1200, 2310, 1540, 1900],
]

GRADE_BANDS = [["grade", "min_base", "max_base"], ["G4", 6000, 9000], ["G5", 8500, 12000], ["G6", 11000, 15000], ["G7", 15000, 21000], ["G8", 21000, 30000]]
DEPARTMENTS = [["department", "owner", "cost_center"], ["HR", "Compensation Team", "CC-HR"], ["Finance", "Finance BP", "CC-FIN"], ["Sales", "Sales Ops", "CC-SAL"], ["Tech", "Tech HRBP", "CC-TEC"], ["Ops", "Operations BP", "CC-OPS"]]
STATUSES = [["status"], ["Active"], ["Inactive"], ["Leave"]]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def add_table_style(ws):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        ws.column_dimensions[letter].width = min(max(width, 12), 34)


def build_workbook(output: Path):
    wb = Workbook()

    raw = wb.active
    raw.title = "raw_payroll_data"
    raw.append(RAW_HEADERS)
    for row in RAW_ROWS:
        raw.append(row)
    style_header(raw)
    add_table_style(raw)

    lookup = wb.create_sheet("lookup_tables")
    for row in GRADE_BANDS:
        lookup.append(row)
    lookup.append([])
    dept_start = lookup.max_row + 1
    for row in DEPARTMENTS:
        lookup.append(row)
    lookup.append([])
    status_start = lookup.max_row + 1
    for row in STATUSES:
        lookup.append(row)
    for row_idx in (1, dept_start, status_start):
        style_header(lookup, row_idx)
    add_table_style(lookup)

    status_range = f"lookup_tables!$A${status_start + 1}:$A${status_start + len(STATUSES) - 1}"
    valid_status = DataValidation(type="list", formula1=f"={status_range}", allow_blank=False)
    raw.add_data_validation(valid_status)
    valid_status.add(f"G2:G{raw.max_row}")

    checks = wb.create_sheet("validation_checks")
    checks.append(RAW_HEADERS + ["gross_pay", "net_pay", "duplicate_check", "missing_required", "grade_min", "grade_max", "band_check", "status_check", "final_check"])
    for idx in range(2, raw.max_row + 1):
        row = [f"=raw_payroll_data!{get_column_letter(col)}{idx}" for col in range(1, len(RAW_HEADERS) + 1)]
        row += [
            f"=H{idx}+I{idx}+J{idx}",
            f"=N{idx}-K{idx}-L{idx}-M{idx}",
            f'=IF(COUNTIFS(B:B,B{idx},A:A,A{idx})>1,"Duplicate","OK")',
            f'=IF(OR(A{idx}="",B{idx}="",D{idx}="",F{idx}="",G{idx}=""),"Missing required field","OK")',
            f'=IFERROR(VLOOKUP(F{idx},lookup_tables!$A$2:$C$6,2,FALSE),"Grade not found")',
            f'=IFERROR(VLOOKUP(F{idx},lookup_tables!$A$2:$C$6,3,FALSE),"Grade not found")',
            f'=IF(OR(NOT(ISNUMBER(Q{idx})),H{idx}<Q{idx},H{idx}>R{idx}),"Out of band","OK")',
            f'=IF(COUNTIF({status_range},G{idx})=1,"OK","Invalid status")',
            f'=IF(OR(O{idx}<>"OK",P{idx}<>"OK",S{idx}<>"OK",T{idx}<>"OK"),"Review","OK")',
        ]
        checks.append(row)
    style_header(checks)
    add_table_style(checks)
    checks.conditional_formatting.add(f"U2:U{checks.max_row}", FormulaRule(formula=['U2="Review"'], fill=PatternFill("solid", fgColor="F4CCCC")))

    summary = wb.create_sheet("summary_report")
    summary.append(["department", "headcount", "gross_pay", "avg_base_salary", "net_pay", "review_rows"])
    for row_idx, dept in enumerate(["HR", "Finance", "Sales", "Tech", "Ops"], start=2):
        summary.append([
            dept,
            f'=COUNTIFS(validation_checks!D:D,A{row_idx},validation_checks!G:G,"Active")',
            f'=SUMIFS(validation_checks!N:N,validation_checks!D:D,A{row_idx})',
            f'=AVERAGEIFS(validation_checks!H:H,validation_checks!D:D,A{row_idx})',
            f'=SUMIFS(validation_checks!O:O,validation_checks!D:D,A{row_idx})',
            f'=COUNTIFS(validation_checks!D:D,A{row_idx},validation_checks!U:U,"Review")',
        ])
    style_header(summary)
    add_table_style(summary)

    notes = wb.create_sheet("process_notes")
    for row in [
        ["topic", "note"],
        ["流程", "保留原始数据，使用字典表统一口径，再用公式生成异常标记和汇总报表。"],
        ["公式", "使用 VLOOKUP、COUNTIFS、SUMIFS、AVERAGEIFS、IF 等公式完成核对和统计。"],
        ["数据质量", "重点检查重复记录、缺失字段、无效状态和职级薪酬区间异常。"],
        ["保密", "共享前隐藏姓名、员工 ID 和明细薪酬，只保留必要统计口径。"],
    ]:
        notes.append(row)
    style_header(notes)
    add_table_style(notes)
    notes.column_dimensions["B"].width = 80
    for cell in notes["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    parser = argparse.ArgumentParser(description="Build a HR compensation Excel demo workbook.")
    parser.add_argument("--output", default="assets/hr_compensation_demo.xlsx", help="Output .xlsx path")
    args = parser.parse_args()
    build_workbook(Path(args.output))
    print(f"Created {args.output}")


if __name__ == "__main__":
    main()
